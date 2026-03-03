"""
Hate Speech Analyser — Desktop App
=========================================
Opens a native desktop window (no browser needed).

Install:
    pip install pywebview transformers torch pandas openpyxl

Run:
    python app.py
"""

import json, re, io, threading
from pathlib import Path
import pandas as pd
import webview
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

VIOLENCE_RE = re.compile(
    r"\b(kill|shoot|murder|bomb|attack|lynch|hang|stab|die|death threat|"
    r"destroy|eliminate|exterminate|execute|slaughter|massacre|hurt|harm|"
    r"beat|punch|rape|assault)\b", re.IGNORECASE)

FILL_COLORS = {"Neutral":"C8E6C9","Offensive":"FFE0B2","Hate":"FFCDD2"}

def detect_format(data):
    if not data: return "old"
    s = data[0]
    return "new" if ("createdAt" in s or "likeCount" in s or "isReply" in s) else "old"

def get_lang(item, fmt):
    if fmt == "new": return item.get("lang","")
    if isinstance(item.get("tweet"),dict): return item["tweet"].get("lang","")
    return item.get("lang","")

def parse_tweet(item, fmt):
    if fmt == "new":
        auth = item.get("author",{}).get("userName","") if isinstance(item.get("author"),dict) else ""
        return {"id":str(item.get("id","")),"created_at":item.get("createdAt",""),"author":auth,
                "text":item.get("text","").strip(),"retweets":item.get("retweetCount",0),
                "likes":item.get("likeCount",0),"lang":item.get("lang",""),"is_reply":item.get("isReply",False)}
    text = item.get("full_text") or item.get("text") or ""
    auth = item["handle"] if item.get("handle") else (item["user"].get("screen_name","") if isinstance(item.get("user"),dict) else item.get("author_id",""))
    lang = item.get("tweet",{}).get("lang","") if isinstance(item.get("tweet"),dict) else item.get("lang","")
    return {"id":str(item.get("id_str") or item.get("id","")),"created_at":item.get("created_at",""),
            "author":auth,"text":text.strip(),"retweets":item.get("retweet_count",0),
            "likes":item.get("favorite_count",0),"lang":lang,
            "is_reply":bool(item.get("tweet",{}).get("in_reply_to_status_id")) if isinstance(item.get("tweet"),dict) else False}

def merge_files(file_paths, fmt):
    seen,tweets = set(),[]
    skipped_lang = skipped_dupe = 0
    per_file = {}
    for path in file_paths:
        with open(path,"r",encoding="utf-8") as f: data = json.load(f)
        added = 0
        for item in data:
            if get_lang(item,fmt) != "en": skipped_lang+=1; continue
            tid = str(item.get("id") or item.get("id_str") or "")
            if tid and tid in seen: skipped_dupe+=1; continue
            seen.add(tid); tweets.append(parse_tweet(item,fmt)); added+=1
        per_file[Path(path).name] = added
    df = pd.DataFrame(tweets)
    df = df[df["text"].str.len()>0].reset_index(drop=True)
    return df, {"per_file":per_file,"skipped_lang":skipped_lang,"skipped_dupe":skipped_dupe}

_clf = None
def get_clf():
    global _clf
    if _clf is None:
        from transformers import pipeline as hfp
        _clf = hfp("text-classification",model="cardiffnlp/twitter-roberta-base-offensive",truncation=True,max_length=512)
    return _clf

def classify_df(df, progress_cb):
    clf = get_clf()
    texts = df["text"].tolist(); total = len(texts)
    ids,names,scores = [],[],[]
    for i,text in enumerate(texts,1):
        try:
            r = clf(text[:512])[0]; lbl = r["label"].lower(); sc = round(r["score"],4)
            if "non" in lbl or lbl=="not-offensive": lid,lname=0,"Neutral"
            elif VIOLENCE_RE.search(text): lid,lname=2,"Hate"
            else: lid,lname=1,"Offensive"
        except: lid,lname,sc=0,"Neutral",0.0
        ids.append(lid); names.append(lname); scores.append(sc)
        if i%20==0 or i==total: progress_cb(i,total)
    df["label_id"]=ids; df["label_name"]=names; df["confidence"]=scores
    return df

def build_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        df.to_excel(w,sheet_name="Results",index=False)
        counts = df["label_name"].value_counts()
        pd.DataFrame({"Category":counts.index,"Count":counts.values,
                      "Percentage":(counts.values/len(df)*100).round(2)}).to_excel(w,sheet_name="Summary",index=False)
    buf.seek(0); wb=load_workbook(buf); ws=wb["Results"]
    col_idx = df.columns.tolist().index("label_name")+1
    for cell in ws[1]:
        cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="37474F")
        cell.alignment=Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        lv=row[col_idx-1].value or "Neutral"
        fill=PatternFill("solid",fgColor=FILL_COLORS.get(lv,"FFFFFF"))
        for cell in row: cell.fill=fill
    for sheet in [ws,wb["Summary"]]:
        for col in sheet.columns:
            ml=max((len(str(c.value or "")) for c in col),default=10)
            sheet.column_dimensions[get_column_letter(col[0].column)].width=min(ml+4,80)
    out=io.BytesIO(); wb.save(out); return out.getvalue()

class API:
    def __init__(self):
        self._progress=0; self._status=""; self._done=False
        self._results=None; self._df=None

    def pick_files(self):
        paths = webview.windows[0].create_file_dialog(
            webview.OPEN_DIALOG,allow_multiple=True,file_types=("JSON Files (*.json)",))
        if not paths: return []
        info=[]
        for p in paths:
            try:
                with open(p,"r",encoding="utf-8") as f: data=json.load(f)
                fmt=detect_format(data)
                info.append({"name":Path(p).name,"path":p,"format":fmt,"count":len(data)})
            except: info.append({"name":Path(p).name,"path":p,"format":"error","count":0})
        return info

    def start_processing(self,file_paths,fmt):
        self._progress=0; self._status="Merging files…"; self._done=False; self._results=None
        t=threading.Thread(target=self._run,args=(file_paths,fmt),daemon=True); t.start()
        return True

    def _run(self,file_paths,fmt):
        try:
            df,stats=merge_files(file_paths,fmt); total=len(df)
            if total==0:
                self._results={"error":"No English tweets found after filtering."}; self._done=True; return
            self._status="Loading model (downloads ~500 MB on first run)…"; self._progress=3
            get_clf()
            self._status=f"Classifying {total:,} tweets…"; self._progress=5
            def cb(i,t): self._progress=5+int((i/t)*90); self._status=f"Classifying tweet {i:,} of {t:,}…"
            df=classify_df(df,cb); self._df=df
            counts=df["label_name"].value_counts().reindex(["Neutral","Offensive","Hate"],fill_value=0)
            pct=(counts/total*100).round(2)
            replies=int(df["is_reply"].sum()) if "is_reply" in df.columns else 0
            self._results={
                "total":total,"counts":counts.to_dict(),"pct":pct.to_dict(),
                "merge_stats":{"per_file":stats["per_file"],"skipped_lang":stats["skipped_lang"],"skipped_dupe":stats["skipped_dupe"]},
                "replies":replies,
                "hate_rows":df[df["label_name"]=="Hate"][["author","text","confidence","likes","retweets"]].head(100).to_dict(orient="records"),
                "off_rows":df[df["label_name"]=="Offensive"][["author","text","confidence","likes","retweets"]].head(100).to_dict(orient="records"),
                "all_rows":df[["author","created_at","text","label_name","confidence","likes","retweets"]].head(300).to_dict(orient="records"),
            }
            self._progress=100; self._status="Done."
        except Exception as e: self._results={"error":str(e)}
        finally: self._done=True

    def get_progress(self): return {"progress":self._progress,"status":self._status,"done":self._done}
    def get_results(self): return self._results

    def minimize(self):
        if webview.windows: webview.windows[0].minimize()
    def maximize(self):
        if webview.windows: webview.windows[0].toggle_fullscreen()
    def close(self):
        if webview.windows: webview.windows[0].destroy()

    def save_excel(self):
        if self._df is None: return False
        path=webview.windows[0].create_file_dialog(webview.SAVE_DIALOG,save_filename="cop28_results.xlsx",file_types=("Excel Files (*.xlsx)",))
        if path:
            sp=path[0] if isinstance(path,(list,tuple)) else path
            with open(sp,"wb") as f: f.write(build_excel(self._df))
            return True
        return False

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Hate Speech Analyser</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#0d1117;--surface:#161b27;--border:#1e2739;--muted:#4a5568;--dim:#2d3748;--text:#e2e8f0;--sub:#8896a5;--green:#22c55e;--orange:#f97316;--red:#ef4444;--blue:#3b82f6;--mono:'DM Mono',monospace;--serif:'DM Serif Display',serif;--sans:'DM Sans',sans-serif}
html,body{width:100vw;height:100vh;overflow:hidden;background:var(--bg);color:var(--text);font-family:var(--sans);font-size:13px;user-select:none}
.shell{display:flex;flex-direction:column;height:100vh}
.topbar{display:flex;align-items:center;gap:10px;padding:0 0 0 20px;background:var(--bg);height:40px;flex-shrink:0;-webkit-app-region:drag;user-select:none;}
.topbar-logo{font-family:var(--serif);font-size:17px;color:var(--text)}.topbar-logo span{color:var(--green)}
.topbar-meta{font-family:var(--mono);font-size:9.5px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-left:auto}
.window-controls{display:flex;height:100%;margin-left:15px;-webkit-app-region:no-drag;}
.win-btn{background:transparent;border:none;color:var(--muted);width:46px;height:100%;display:flex;align-items:center;justify-content:center;cursor:pointer;transition:all .2s;}
.win-btn:hover{background:rgba(255,255,255,.1);color:var(--text);}
.win-btn.close:hover{background:#ef4444;color:#fff;}
.body{display:flex;flex:1;overflow:hidden}
.left{width:295px;flex-shrink:0;border-right:1px solid var(--border);background:var(--surface);display:flex;flex-direction:column;padding:14px;gap:12px;overflow-y:auto;overflow-x:hidden}
.lbl{font-family:var(--mono);font-size:9.5px;color:var(--muted);text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px}
.upload-zone{border:1.5px dashed var(--dim);border-radius:10px;padding:14px 10px;text-align:center;cursor:pointer;transition:border-color .2s,background .2s;background:rgba(255,255,255,.02)}
.upload-zone:hover{border-color:var(--green);background:rgba(34,197,94,.04)}
.upload-zone .ico{font-size:20px;margin-bottom:5px}.upload-zone p{font-size:11px;color:var(--sub)}.upload-zone p strong{color:var(--green)}
.file-list{display:flex;flex-direction:column;gap:4px;max-height:130px;overflow-y:auto}
.fi{display:flex;align-items:center;gap:5px;background:rgba(255,255,255,.03);border:1px solid var(--border);border-radius:7px;padding:5px 8px;font-size:10.5px}
.fi .fn{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--sub)}
.badge{font-family:var(--mono);font-size:9px;font-weight:500;padding:2px 6px;border-radius:999px;white-space:nowrap}
.bn{background:rgba(59,130,246,.15);color:#60a5fa;border:1px solid rgba(59,130,246,.25)}
.bo{background:rgba(168,85,247,.15);color:#c084fc;border:1px solid rgba(168,85,247,.25)}
.warn{background:rgba(251,191,36,.08);border:1px solid rgba(251,191,36,.25);border-radius:8px;padding:7px 10px;font-size:10.5px;color:#fbbf24;font-family:var(--mono);display:none}
select{width:100%;background:var(--bg);border:1px solid var(--border);color:var(--sub);border-radius:7px;padding:7px 10px;font-family:var(--mono);font-size:10.5px;cursor:pointer;appearance:none;outline:none}
select:focus{border-color:var(--green)}
.ms{display:grid;grid-template-columns:1fr 1fr;gap:5px}
.msc{background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:7px 8px;text-align:center}
.msn{font-family:var(--serif);font-size:16px;line-height:1}.msl{font-family:var(--mono);font-size:8.5px;color:var(--muted);text-transform:uppercase;margin-top:2px}
.btn-run{background:var(--green);color:#0d1117;border:none;border-radius:8px;padding:9px;font-family:var(--sans);font-weight:600;font-size:13px;cursor:pointer;transition:background .2s,transform .1s;margin-top:auto;width:100%}
.btn-run:hover:not(:disabled){background:#16a34a;transform:translateY(-1px)}.btn-run:disabled{background:var(--dim);color:var(--muted);cursor:not-allowed;transform:none}
.prog-wrap{display:none;flex-direction:column;gap:4px}
.prog-bg{height:4px;background:var(--dim);border-radius:99px;overflow:hidden}
.prog-fill{height:100%;background:var(--green);border-radius:99px;transition:width .3s ease;width:0%}
.prog-txt{font-family:var(--mono);font-size:9.5px;color:var(--muted)}
.right{flex:1;display:flex;flex-direction:column;padding:14px;gap:10px;overflow:hidden}
.placeholder{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:8px;color:var(--dim)}
.placeholder .big{font-family:var(--serif);font-size:26px;color:var(--border)}.placeholder p{font-family:var(--mono);font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.stat-row{display:grid;grid-template-columns:repeat(4,1fr);gap:7px;flex-shrink:0}
.sc{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:9px 12px;text-align:center}
.sn{font-family:var(--serif);font-size:21px;line-height:1.1}.sl{font-family:var(--mono);font-size:8.5px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-top:2px}
.charts-row{display:grid;grid-template-columns:200px 1fr 1fr;gap:8px;flex-shrink:0;height:240px}
.cc{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:9px 12px;display:flex;flex-direction:column}
.cc canvas{flex:1;min-height:0}
.tabs-bar{display:flex;gap:2px;flex-shrink:0}
.tb{font-family:var(--mono);font-size:9.5px;text-transform:uppercase;letter-spacing:.07em;padding:5px 11px;border-radius:6px;border:1px solid transparent;background:none;color:var(--muted);cursor:pointer;transition:all .15s}
.tb:hover{color:var(--text);background:var(--border)}.tb.active{background:var(--surface);border-color:var(--border);color:var(--text)}
.tbl-wrap{flex:1;overflow-y:auto;overflow-x:hidden;background:var(--surface);border:1px solid var(--border);border-radius:10px;min-height:0}
table{width:100%;border-collapse:collapse}
thead th{position:sticky;top:0;background:var(--bg);font-family:var(--mono);font-size:9px;text-transform:uppercase;letter-spacing:.07em;color:var(--muted);padding:7px 9px;border-bottom:1px solid var(--border);text-align:left}
tbody tr{border-bottom:1px solid rgba(255,255,255,.03)}tbody tr:hover{background:rgba(255,255,255,.02)}
tbody td{padding:5px 9px;font-size:11px;color:var(--sub);vertical-align:top}
td.tc{max-width:400px;line-height:1.4;color:var(--text);word-wrap:break-word}
.pill{display:inline-block;font-family:var(--mono);font-size:8.5px;padding:2px 6px;border-radius:999px;font-weight:500}
.pN{background:rgba(34,197,94,.12);color:#4ade80}.pO{background:rgba(249,115,22,.12);color:#fb923c}.pH{background:rgba(239,68,68,.12);color:#f87171}
.bot{display:flex;align-items:center;gap:8px;flex-shrink:0}
.btn-dl{background:var(--surface);color:var(--sub);border:1px solid var(--border);border-radius:8px;padding:6px 14px;font-family:var(--mono);font-size:10px;cursor:pointer;transition:all .15s}
.btn-dl:hover{border-color:var(--green);color:var(--green)}
.ri{font-family:var(--mono);font-size:9.5px;color:var(--muted);margin-left:auto}
::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:transparent}::-webkit-scrollbar-thumb{background:var(--dim);border-radius:99px}
</style>
</head>
<body>
<div class="shell">
<div class="topbar pywebview-drag-region">
  <div class="topbar-logo"><span>Hate Speech</span> Analyser</div>
  <div class="topbar-meta">TU DORTMUND · LANGUAGE IN THE MEDIA · WISE 2023/2024</div>
  <div class="window-controls">
    <button class="win-btn" onclick="pywebview.api.minimize()" title="Minimize">
      <svg width="12" height="12" viewBox="0 0 12 12"><rect fill="currentColor" width="10" height="1" x="1" y="6"></rect></svg>
    </button>
    <button class="win-btn" onclick="pywebview.api.maximize()" title="Maximize">
      <svg width="12" height="12" viewBox="0 0 12 12"><rect stroke="currentColor" fill="none" width="10" height="10" x="1" y="1"></rect></svg>
    </button>
    <button class="win-btn close" onclick="pywebview.api.close()" title="Close">
      <svg width="12" height="12" viewBox="0 0 12 12"><polygon fill="currentColor" points="11,1.5 10.5,1 6,5.5 1.5,1 1,1.5 5.5,6 1,10.5 1.5,11 6,6.5 10.5,11 11,10.5 6.5,6 "></polygon></svg>
    </button>
  </div>
</div>
<div class="body">
<div class="left">
  <div>
    <div class="lbl">① Upload Datasets</div>
    <div class="upload-zone" onclick="pickFiles()">
      <div class="ico">📂</div>
      <p><strong>Click to add files</strong></p>
      <p>Apify JSON exports (.json)</p>
    </div>
  </div>
  <div id="flWrap" style="display:none">
    <div class="lbl">Files</div>
    <div class="file-list" id="fileList"></div>
  </div>
  <div id="warnEl" class="warn">⚠ Mixed formats detected — files must be from the same actor. Remove mismatched files or select a format manually.</div>
  <div>
    <div class="lbl">② Actor Format</div>
    <select id="fmtSel">
      <option value="auto">🔍 Auto-detect</option>
      <option value="new">apidojo / twitter-scraper-lite</option>
      <option value="old">altimis / scweet</option>
    </select>
  </div>
  <div id="msWrap" style="display:none">
    <div class="lbl">Merge Result</div>
    <div class="ms" id="msDiv"></div>
  </div>
  <div id="progWrap" class="prog-wrap">
    <div class="prog-bg"><div class="prog-fill" id="pFill"></div></div>
    <div class="prog-txt" id="pTxt">Starting…</div>
  </div>
  <button class="btn-run" id="runBtn" disabled onclick="run()">▶ Run Analysis</button>
</div>
<div class="right" id="rightPanel">
  <div class="placeholder" id="ph">
    <div class="big">🌍</div>
    <p>Upload JSON files to begin</p>
  </div>
  <div id="res" style="display:none;flex-direction:column;gap:10px;height:100%;overflow:hidden">
    <div class="stat-row">
      <div class="sc"><div class="sn" id="sT" style="color:var(--text)">—</div><div class="sl">Total Tweets</div></div>
      <div class="sc"><div class="sn" id="sN" style="color:var(--green)">—</div><div class="sl">Neutral</div></div>
      <div class="sc"><div class="sn" id="sO" style="color:var(--orange)">—</div><div class="sl">Offensive</div></div>
      <div class="sc"><div class="sn" id="sH" style="color:var(--red)">—</div><div class="sl">Hate Speech</div></div>
    </div>
    <div class="charts-row">
      <div class="cc"><div class="lbl">Distribution</div><div style="position:relative;flex:1;min-height:0;overflow:hidden"><canvas id="pieC" style="position:absolute;top:0;left:0;width:100%;height:100%"></canvas></div></div>
      <div class="cc"><div class="lbl">Breakdown</div><div style="position:relative;flex:1;min-height:0;overflow:hidden"><canvas id="barC" style="position:absolute;top:0;left:0;width:100%;height:100%"></canvas></div></div>
      <div class="cc"><div class="lbl">vs. Original Study (2024)</div><div style="position:relative;flex:1;min-height:0;overflow:hidden"><canvas id="cmpC" style="position:absolute;top:0;left:0;width:100%;height:100%"></canvas></div></div>
    </div>
    <div class="tabs-bar">
      <button class="tb active" onclick="tab('hate')" id="t-hate">🔴 Hate</button>
      <button class="tb" onclick="tab('off')" id="t-off">🟠 Offensive</button>
      <button class="tb" onclick="tab('all')" id="t-all">📋 All Results</button>
    </div>
    <div class="tbl-wrap" id="tblWrap"></div>
    <div class="bot">
      <button class="btn-dl" onclick="saveExcel()">⬇ Save Excel</button>
      <div class="ri" id="ri"></div>
    </div>
  </div>
</div>
</div>
</div>
<script>
let files=[],pInt=null,pC=null,bC=null,cC=null,tData={hate:[],off:[],all:[]};
const Co={Neutral:'#22c55e',Offensive:'#f97316',Hate:'#ef4444'};
const CHART_BASE={responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}}};

async function pickFiles(){
  const picked=await pywebview.api.pick_files();
  if(!picked||!picked.length)return;
  const ex=new Set(files.map(f=>f.path));
  picked.forEach(f=>{if(!ex.has(f.path))files.push(f);});
  renderFiles();checkMixed();
}
function renderFiles(){
  const list=document.getElementById('fileList'),wrap=document.getElementById('flWrap');
  list.innerHTML='';
  if(!files.length){wrap.style.display='none';return;}
  wrap.style.display='block';
  files.forEach((f,i)=>{
    const bc=f.format==='new'?'bn':'bo',bl=f.format==='new'?'scraper-lite':'scweet';
    list.innerHTML+=`<div class="fi"><span class="fn" title="${f.name}">${f.name}</span><span class="badge ${bc}">${bl}</span><span style="color:var(--muted);font-size:10px;cursor:pointer" onclick="rm(${i})">✕</span></div>`;
  });
  document.getElementById('runBtn').disabled=!files.length;
}
function rm(i){files.splice(i,1);renderFiles();checkMixed();if(!files.length)document.getElementById('runBtn').disabled=true;}
function checkMixed(){
  const w=document.getElementById('warnEl'),fmt=document.getElementById('fmtSel').value;
  if(fmt!=='auto'){w.style.display='none';return;}
  const fmts=[...new Set(files.map(f=>f.format).filter(f=>f!=='error'))];
  w.style.display=fmts.length>1?'block':'none';
}
document.getElementById('fmtSel').addEventListener('change',checkMixed);

async function run(){
  if(!files.length)return;
  let fmt=document.getElementById('fmtSel').value;
  if(fmt==='auto'){
    const fmts=[...new Set(files.map(f=>f.format).filter(f=>f!=='error'))];
    if(fmts.length>1)return;
    fmt=fmts[0]||'old';
  }
  document.getElementById('runBtn').disabled=true;
  document.getElementById('progWrap').style.display='flex';
  document.getElementById('msWrap').style.display='none';
  document.getElementById('ph').style.display='flex';
  document.getElementById('res').style.display='none';
  await pywebview.api.start_processing(files.map(f=>f.path),fmt);
  pInt=setInterval(async()=>{
    const p=await pywebview.api.get_progress();
    document.getElementById('pFill').style.width=p.progress+'%';
    document.getElementById('pTxt').textContent=p.status;
    if(p.done){
      clearInterval(pInt);
      const r=await pywebview.api.get_results();
      if(r&&r.error){document.getElementById('pTxt').textContent='⚠ '+r.error;document.getElementById('runBtn').disabled=false;}
      else if(r)render(r);
    }
  },400);
}

function render(r){
  document.getElementById('progWrap').style.display='none';
  document.getElementById('ph').style.display='none';
  document.getElementById('res').style.display='flex';
  document.getElementById('runBtn').disabled=false;
  const ms=document.getElementById('msDiv');
  document.getElementById('msWrap').style.display='block';
  let mh='';
  Object.entries(r.merge_stats.per_file).forEach(([n,c])=>{mh+=`<div class="msc"><div class="msn" style="color:var(--blue)">${c.toLocaleString()}</div><div class="msl">${n.substring(0,12)}</div></div>`;});
  mh+=`<div class="msc"><div class="msn" style="color:var(--orange)">${r.merge_stats.skipped_lang.toLocaleString()}</div><div class="msl">Non-English</div></div>`;
  mh+=`<div class="msc"><div class="msn" style="color:var(--muted)">${r.merge_stats.skipped_dupe.toLocaleString()}</div><div class="msl">Duplicates</div></div>`;
  ms.innerHTML=mh;
  document.getElementById('sT').textContent=r.total.toLocaleString();
  document.getElementById('sN').textContent=r.pct['Neutral'].toFixed(1)+'%';
  document.getElementById('sO').textContent=r.pct['Offensive'].toFixed(1)+'%';
  document.getElementById('sH').textContent=r.pct['Hate'].toFixed(1)+'%';
  document.getElementById('ri').textContent=`${r.replies.toLocaleString()} replies (${(r.replies/r.total*100).toFixed(1)}%)`;
  drawCharts(r);
  tData={hate:r.hate_rows,off:r.off_rows,all:r.all_rows};
  tab('hate');
}

function drawCharts(r){
  if(pC)pC.destroy();if(bC)bC.destroy();if(cC)cC.destroy();
  const cats=['Neutral','Offensive','Hate'];
  pC=new Chart(document.getElementById('pieC'),{type:'doughnut',data:{labels:cats,datasets:[{data:cats.map(c=>r.counts[c]),backgroundColor:cats.map(c=>Co[c]),borderColor:'#0d1117',borderWidth:3,hoverOffset:5}]},options:{...CHART_BASE,cutout:'62%',plugins:{legend:{display:true,position:'bottom',labels:{color:'#8896a5',font:{family:'DM Mono',size:8.5},boxWidth:9,padding:6}},tooltip:{callbacks:{label:ctx=>` ${ctx.label}: ${ctx.parsed.toLocaleString()} (${r.pct[ctx.label].toFixed(2)}%)`}}}}});
  bC=new Chart(document.getElementById('barC'),{type:'bar',data:{labels:cats,datasets:[{data:cats.map(c=>r.pct[c]),backgroundColor:cats.map(c=>Co[c]),borderColor:'#0d1117',borderWidth:2,borderRadius:5}]},options:{...CHART_BASE,scales:{x:{grid:{display:false},ticks:{color:'#4a5568',font:{family:'DM Mono',size:9}}},y:{grid:{color:'rgba(255,255,255,.04)'},ticks:{color:'#4a5568',font:{family:'DM Mono',size:9},callback:v=>v+'%'}}},plugins:{...CHART_BASE.plugins,tooltip:{callbacks:{label:ctx=>` ${ctx.parsed.y.toFixed(2)}%`}}}}});
  cC=new Chart(document.getElementById('cmpC'),{type:'bar',data:{labels:cats,datasets:[{label:'Original (2024)',data:[88.54,10.54,0.92],backgroundColor:'rgba(99,102,241,.65)',borderColor:'#0d1117',borderWidth:2,borderRadius:4},{label:'Recreation (2026)',data:cats.map(c=>r.pct[c]),backgroundColor:'rgba(34,197,94,.65)',borderColor:'#0d1117',borderWidth:2,borderRadius:4}]},options:{...CHART_BASE,barPercentage:0.7,scales:{x:{grid:{display:false},ticks:{color:'#4a5568',font:{family:'DM Mono',size:9}}},y:{grid:{color:'rgba(255,255,255,.04)'},ticks:{color:'#4a5568',font:{family:'DM Mono',size:9},callback:v=>v+'%'}}},plugins:{legend:{display:true,position:'bottom',labels:{color:'#8896a5',font:{family:'DM Mono',size:8.5},boxWidth:9,padding:6}},tooltip:{callbacks:{label:ctx=>` ${ctx.dataset.label}: ${ctx.parsed.y.toFixed(2)}%`}}}}});
}

function tab(t){
  ['hate','off','all'].forEach(x=>document.getElementById('t-'+x).classList.toggle('active',x===t));
  const rows=tData[t]||[],isAll=t==='all';
  let h='<table><thead><tr><th>Author</th>';
  if(isAll)h+='<th>Date</th>';
  h+='<th>Text</th>';
  if(isAll)h+='<th>Label</th>';
  h+='<th>Conf.</th><th>Likes</th>';
  if(!isAll)h+='<th>RTs</th>';
  h+='</tr></thead><tbody>';
  rows.forEach(r=>{
    h+=`<tr><td style="white-space:nowrap">@${esc(r.author||'—')}</td>`;
    if(isAll)h+=`<td style="white-space:nowrap;color:var(--muted);font-size:10px">${esc((r.created_at||'').substring(4,16))}</td>`;
    h+=`<td class="tc" title="${esc(r.text)}">${esc(r.text)}</td>`;
    if(isAll)h+=`<td><span class="pill p${r.label_name[0]}">${esc(r.label_name)}</span></td>`;
    h+=`<td style="font-family:var(--mono);font-size:10px">${(r.confidence||0).toFixed(2)}</td>`;
    h+=`<td style="font-family:var(--mono);font-size:10px">${(r.likes||0).toLocaleString()}</td>`;
    if(!isAll)h+=`<td style="font-family:var(--mono);font-size:10px">${(r.retweets||0).toLocaleString()}</td>`;
    h+='</tr>';
  });
  h+='</tbody></table>';
  document.getElementById('tblWrap').innerHTML=h;
}

function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');}
async function saveExcel(){await pywebview.api.save_excel();}
</script>
</body>
</html>"""

if __name__ == "__main__":
    api = API()
    window = webview.create_window(
        title="Hate Speech Analyser",
        html=HTML, js_api=api,
        width=1280, height=780,
        resizable=True, min_size=(1280, 780),
        frameless=True, easy_drag=False
    )
    webview.start(debug=False)
